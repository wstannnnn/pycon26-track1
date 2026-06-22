"""Join SkillsFuture Skills Framework workbook tabs into role-level records.

The source workbook is denormalized across tabs:
- Job Role_Description: role summary and performance expectation
- Job Role_CWF_KT: role critical work functions and key tasks
- Job Role_TCS_CCS: role-to-skill/proficiency mappings
- TSC_CCS_Key and TSC_CCS_Key_Retired: active/retired skill definitions
- TSC_CCS_K&A: proficiency descriptions plus knowledge/ability items

This script reads all tabs and writes one joined record per job role. Each
record includes structured tasks/skills and a concatenated_description field
that is useful for search indexing or LLM retrieval.
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_WORKBOOK = REPO_ROOT / "data" / "jobsandskills-skillsfuture-skills-framework-dataset.xlsx"
DEFAULT_OUTPUT = REPO_ROOT / "data" / "processed" / "skills_framework_joined_roles.jsonl"

SHEET_ROLE_DESCRIPTION = "Job Role_Description"
SHEET_ROLE_TASKS = "Job Role_CWF_KT"
SHEET_ROLE_SKILLS = "Job Role_TCS_CCS"
SHEET_SKILL_KEY = "TSC_CCS_Key"
SHEET_SKILL_KEY_RETIRED = "TSC_CCS_Key_Retired"
SHEET_SKILL_KA = "TSC_CCS_K&A"


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    output_path = args.output.resolve()

    joined_roles = build_joined_roles(workbook_path)
    write_records(joined_roles, output_path, args.format)

    print(f"Read workbook: {workbook_path}")
    print(f"Wrote {len(joined_roles)} joined role records: {output_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Join all tabs in the SkillsFuture Skills Framework workbook."
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"Path to source .xlsx workbook. Default: {DEFAULT_WORKBOOK}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Path to write the joined output. Default: {DEFAULT_OUTPUT}",
    )
    parser.add_argument(
        "--format",
        choices=("jsonl", "csv"),
        default="jsonl",
        help="Output format. JSONL preserves nested tasks and skills; CSV flattens them as JSON strings.",
    )
    return parser.parse_args()


def build_joined_roles(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    validate_sheets(workbook.sheetnames)

    role_descriptions = read_sheet(workbook[SHEET_ROLE_DESCRIPTION])
    role_tasks = read_sheet(workbook[SHEET_ROLE_TASKS])
    role_skills = read_sheet(workbook[SHEET_ROLE_SKILLS])
    active_skill_keys = read_sheet(workbook[SHEET_SKILL_KEY])
    retired_skill_keys = read_sheet(workbook[SHEET_SKILL_KEY_RETIRED])
    skill_knowledge_abilities = read_sheet(workbook[SHEET_SKILL_KA])

    tasks_by_role = group_tasks_by_role(role_tasks)
    skill_defs_by_code = build_skill_definitions(active_skill_keys, retired_skill_keys)
    knowledge_by_skill_level = group_knowledge_abilities(skill_knowledge_abilities)
    skills_by_role = group_skills_by_role(role_skills, skill_defs_by_code, knowledge_by_skill_level)

    joined_roles: list[dict[str, Any]] = []
    seen_roles: set[tuple[str, str, str]] = set()

    for row in role_descriptions:
        role_key = make_role_key(row)
        if not role_key[2] or role_key in seen_roles:
            continue

        seen_roles.add(role_key)
        tasks = tasks_by_role.get(role_key, [])
        skills = skills_by_role.get(role_key, [])

        role_record = {
            "sector": clean(row.get("Sector")),
            "track": clean(row.get("Track")),
            "job_role": clean(row.get("Job Role")),
            "job_role_description": clean(row.get("Job Role Description")),
            "performance_expectation": clean(row.get("Performance Expectation")),
            "critical_work_functions": tasks,
            "skills": skills,
        }
        role_record["concatenated_description"] = build_concatenated_description(role_record)
        joined_roles.append(role_record)

    return joined_roles


def validate_sheets(sheet_names: list[str]) -> None:
    required = {
        SHEET_ROLE_DESCRIPTION,
        SHEET_ROLE_TASKS,
        SHEET_ROLE_SKILLS,
        SHEET_SKILL_KEY,
        SHEET_SKILL_KEY_RETIRED,
        SHEET_SKILL_KA,
    }
    missing = sorted(required.difference(sheet_names))
    if missing:
        raise ValueError(f"Workbook is missing required sheet(s): {', '.join(missing)}")


def read_sheet(worksheet) -> list[dict[str, str]]:
    rows = worksheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    records: list[dict[str, str]] = []

    for row in rows:
        record = {
            header: clean(row[index] if index < len(row) else "")
            for index, header in enumerate(headers)
        }
        if any(record.values()):
            records.append(record)

    return records


def group_tasks_by_role(rows: list[dict[str, str]]) -> dict[tuple[str, str, str], list[dict[str, Any]]]:
    tasks_by_function: dict[tuple[str, str, str], dict[str, list[str]]] = defaultdict(
        lambda: defaultdict(list)
    )

    for row in rows:
        role_key = make_role_key(row)
        function = clean(row.get("Critical Work Function"))
        task = clean(row.get("Key Tasks"))
        if not role_key[2] or not function or not task:
            continue
        append_unique(tasks_by_function[role_key][function], task)

    return {
        role_key: [
            {"critical_work_function": function, "key_tasks": tasks}
            for function, tasks in functions.items()
        ]
        for role_key, functions in tasks_by_function.items()
    }


def build_skill_definitions(
    active_rows: list[dict[str, str]],
    retired_rows: list[dict[str, str]],
) -> dict[str, dict[str, str]]:
    definitions: dict[str, dict[str, str]] = {}

    for row in active_rows:
        code = clean(row.get("TSC Code"))
        if code:
            definitions[code] = {
                "skill_code": code,
                "sector": clean(row.get("Sector")),
                "category": clean(row.get("TSC_CCS Category")),
                "title": clean(row.get("TSC_CCS Title")),
                "description": clean(row.get("TSC_CCS Description")),
                "type": clean(row.get("TSC_CCS Type")),
                "latest_update_date": clean(row.get("Latest Update Date")),
                "retired_date": "",
                "status": "active",
            }

    for row in retired_rows:
        code = clean(row.get("TSC Code"))
        if code:
            definitions[code] = {
                "skill_code": code,
                "sector": clean(row.get("Sector")),
                "category": clean(row.get("TSC_CCS Category")),
                "title": clean(row.get("TSC_CCS Title")),
                "description": clean(row.get("TSC_CCS Description")),
                "type": clean(row.get("TSC_CCS Type")),
                "latest_update_date": "",
                "retired_date": clean(row.get("Retired Date")),
                "status": "retired",
            }

    return definitions


def group_knowledge_abilities(
    rows: list[dict[str, str]],
) -> dict[tuple[str, str], dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}

    for row in rows:
        skill_code = clean(row.get("TSC_CCS Code"))
        proficiency_level = clean(row.get("Proficiency Level"))
        if not skill_code:
            continue

        key = (skill_code, proficiency_level)
        details = grouped.setdefault(
            key,
            {
                "proficiency_description": clean(row.get("Proficiency Description")),
                "knowledge_items": [],
                "ability_items": [],
            },
        )

        if not details["proficiency_description"]:
            details["proficiency_description"] = clean(row.get("Proficiency Description"))

        item = clean(row.get("Knowledge / Ability Items"))
        classification = clean(row.get("Knowledge / Ability Classification")).lower()
        if classification == "knowledge":
            append_unique(details["knowledge_items"], item)
        elif classification == "ability":
            append_unique(details["ability_items"], item)

    return grouped


def group_skills_by_role(
    role_skill_rows: list[dict[str, str]],
    skill_defs_by_code: dict[str, dict[str, str]],
    knowledge_by_skill_level: dict[tuple[str, str], dict[str, Any]],
) -> dict[tuple[str, str, str], list[dict[str, Any]]]:
    grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    seen: set[tuple[str, str, str, str, str]] = set()

    for row in role_skill_rows:
        role_key = make_role_key(row)
        skill_code = clean(row.get("TSC_CCS Code"))
        proficiency_level = clean(row.get("Proficiency Level"))
        if not role_key[2] or not skill_code:
            continue

        dedupe_key = (*role_key, skill_code, proficiency_level)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)

        skill_def = skill_defs_by_code.get(skill_code, {})
        knowledge = knowledge_by_skill_level.get((skill_code, proficiency_level), {})
        title = clean(row.get("TSC_CCS Title")) or clean(skill_def.get("title"))
        skill_type = clean(row.get("TSC_CCS Type")) or clean(skill_def.get("type"))

        grouped[role_key].append(
            {
                "skill_code": skill_code,
                "title": title,
                "type": skill_type,
                "category": clean(skill_def.get("category")),
                "description": clean(skill_def.get("description")),
                "proficiency_level": proficiency_level,
                "proficiency_description": clean(knowledge.get("proficiency_description")),
                "knowledge_items": knowledge.get("knowledge_items", []),
                "ability_items": knowledge.get("ability_items", []),
                "status": clean(skill_def.get("status")) or "unknown",
                "latest_update_date": clean(skill_def.get("latest_update_date")),
                "retired_date": clean(skill_def.get("retired_date")),
            }
        )

    return grouped


def build_concatenated_description(role_record: dict[str, Any]) -> str:
    parts = [
        f"Sector: {role_record['sector']}",
        f"Track: {role_record['track']}",
        f"Job role: {role_record['job_role']}",
        f"Role description: {role_record['job_role_description']}",
        f"Performance expectation: {role_record['performance_expectation']}",
    ]

    for function in role_record["critical_work_functions"]:
        tasks = "; ".join(function["key_tasks"])
        parts.append(f"Critical work function: {function['critical_work_function']}. Key tasks: {tasks}")

    for skill in role_record["skills"]:
        skill_parts = [
            f"Skill: {skill['title']}",
            f"Code: {skill['skill_code']}",
            f"Type: {skill['type']}",
            f"Category: {skill['category']}",
            f"Status: {skill['status']}",
            f"Proficiency level: {skill['proficiency_level']}",
            f"Description: {skill['description']}",
            f"Proficiency description: {skill['proficiency_description']}",
            f"Knowledge: {'; '.join(skill['knowledge_items'])}",
            f"Abilities: {'; '.join(skill['ability_items'])}",
        ]
        parts.append(". ".join(part for part in skill_parts if not part.endswith(": ")))

    return "\n".join(part for part in parts if not part.endswith(": "))


def write_records(records: list[dict[str, Any]], output_path: Path, output_format: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_format == "jsonl":
        with output_path.open("w", encoding="utf-8") as file:
            for record in records:
                file.write(json.dumps(record, ensure_ascii=True) + "\n")
        return

    fieldnames = [
        "sector",
        "track",
        "job_role",
        "job_role_description",
        "performance_expectation",
        "critical_work_functions",
        "skills",
        "concatenated_description",
    ]
    with output_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            row = dict(record)
            row["critical_work_functions"] = json.dumps(
                row["critical_work_functions"], ensure_ascii=True
            )
            row["skills"] = json.dumps(row["skills"], ensure_ascii=True)
            writer.writerow(row)


def make_role_key(row: dict[str, str]) -> tuple[str, str, str]:
    return (clean(row.get("Sector")), clean(row.get("Track")), clean(row.get("Job Role")))


def append_unique(items: list[str], value: str) -> None:
    if value and value not in items:
        items.append(value)


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


if __name__ == "__main__":
    main()

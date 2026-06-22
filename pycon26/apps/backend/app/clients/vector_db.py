import json
from hashlib import sha256
from pathlib import Path

import chromadb
from chromadb.api.models.Collection import Collection
from openpyxl import load_workbook

from app.config import settings
from app.schemas.vectors import VectorPoint, VectorSearchHit


EMBEDDING_DIMENSIONS = 64
BATCH_SIZE = 500
BACKEND_ROOT = Path(__file__).resolve().parents[2]


class VectorDbClient:
    def __init__(
        self,
        path: str = settings.vector_db_path,
        collection: str = settings.vector_db_collection,
        data_dir: str = settings.skills_data_dir,
        auto_index: bool = settings.vector_db_auto_index,
    ) -> None:
        self.path = resolve_backend_path(path)
        self.collection_name = collection
        self.data_dir = resolve_backend_path(data_dir)
        self.auto_index = auto_index
        self._client = chromadb.PersistentClient(path=str(self.path))
        self._collection: Collection | None = None

    @property
    def collection(self) -> Collection:
        if self._collection is None:
            self._collection = self._client.get_or_create_collection(
                name=self.collection_name,
                metadata={"description": "SkillsFuture skills, roles, tasks, and mappings"},
            )
        return self._collection

    def upsert_points(self, points: list[VectorPoint]) -> dict[str, object]:
        documents = [point.document or metadata_to_document(point.payload) for point in points]
        self.collection.upsert(
            ids=[str(point.id) for point in points],
            embeddings=[point.vector or embed_text(doc) for point, doc in zip(points, documents)],
            documents=documents,
            metadatas=[sanitize_metadata(point.payload) for point in points],
        )
        return {"status": "ok", "result": {"upserted": len(points)}}

    def search_text(
        self,
        text: str,
        limit: int = 5,
        where: dict[str, object] | None = None,
    ) -> list[VectorSearchHit]:
        self.ensure_indexed()
        return self.search(vector=embed_text(text), limit=limit, where=where)

    def search(
        self,
        vector: list[float],
        limit: int = 5,
        where: dict[str, object] | None = None,
    ) -> list[VectorSearchHit]:
        query_limit = max(limit * 4, limit)
        query_kwargs = {
            "query_embeddings": [vector],
            "n_results": query_limit,
            "include": ["documents", "metadatas", "distances"],
        }
        if where:
            query_kwargs["where"] = where

        result = self.collection.query(
            **query_kwargs,
        )
        ids = result.get("ids", [[]])[0]
        distances = result.get("distances", [[]])[0]
        metadatas = result.get("metadatas", [[]])[0]
        documents = result.get("documents", [[]])[0]

        matches: list[VectorSearchHit] = []
        for index, item_id in enumerate(ids):
            metadata = dict(metadatas[index] or {})
            metadata["document"] = documents[index]
            distance = float(distances[index] or 0)
            matches.append(
                VectorSearchHit(
                    id=str(item_id),
                    score=round(1 / (1 + distance), 6),
                    payload=metadata,
                )
            )
        return fuse_matches_by_name(matches, limit=limit)

    def find_role_records(self, role_query: str, limit: int = 5) -> list[VectorSearchHit]:
        self.ensure_indexed()
        normalized_query = normalize_tokens(role_query)
        if not normalized_query:
            return []

        result = self.collection.get(
            where={"record_type": "job_role"},
            include=["documents", "metadatas"],
        )
        ids = result.get("ids", [])
        metadatas = result.get("metadatas", [])
        documents = result.get("documents", [])

        matches: list[VectorSearchHit] = []
        for index, item_id in enumerate(ids):
            metadata = dict(metadatas[index] or {})
            role = clean(metadata.get("role"))
            score = role_match_score(normalized_query, role)
            if score <= 0:
                continue

            metadata["document"] = documents[index]
            matches.append(
                VectorSearchHit(
                    id=str(item_id),
                    score=score,
                    payload=metadata,
                )
            )

        return fuse_matches_by_name(matches, limit=limit)

    def ensure_indexed(self) -> None:
        if not self.auto_index or self.collection.count() > 0:
            return
        self.index_data_dir(self.data_dir)

    def index_data_dir(self, data_dir: Path) -> int:
        self.reset_collection()
        records = list(load_skill_records(data_dir))
        for start in range(0, len(records), BATCH_SIZE):
            self.upsert_points(records[start : start + BATCH_SIZE])
        return len(records)

    def reset_collection(self) -> None:
        try:
            self._client.delete_collection(name=self.collection_name)
        except Exception:
            pass
        self._collection = None


def embed_text(text: str, dimensions: int = EMBEDDING_DIMENSIONS) -> list[float]:
    vector = [0.0] * dimensions
    tokens = [token.lower() for token in text.split() if token.strip()]
    if not tokens:
        return vector
    for token in tokens:
        bucket = int(sha256(token.encode("utf-8")).hexdigest(), 16) % dimensions
        vector[bucket] += 1.0
    length = sum(value * value for value in vector) ** 0.5 or 1.0
    return [round(value / length, 6) for value in vector]


def load_skill_records(data_dir: Path):
    yield from load_joined_role_records(
        data_dir / "processed" / "skills_framework_joined_roles.jsonl"
    )


def load_joined_role_records(path: Path):
    if not path.exists():
        return

    with path.open(encoding="utf-8") as file:
        for row_index, line in enumerate(file, start=1):
            if not line.strip():
                continue

            record = json.loads(line)
            role = clean(record.get("job_role"))
            if not role:
                continue

            sector = clean(record.get("sector"))
            track = clean(record.get("track"))
            description = clean(record.get("job_role_description"))
            source = path.name
            yield VectorPoint(
                id=f"joined-role-{row_index}",
                document=clean(record.get("concatenated_description"))
                or f"Role: {role}. Sector: {sector}. Track: {track}. Description: {description}",
                payload={
                    "source": source,
                    "record_type": "job_role",
                    "role": role,
                    "sector": sector,
                    "track": track,
                    "description": description,
                    "performance_expectation": clean(record.get("performance_expectation")),
                },
            )

            for skill_index, skill in enumerate(record.get("skills") or [], start=1):
                skill_title = clean(skill.get("title"))
                if not skill_title:
                    continue
                document = (
                    f"Role skill: {role} requires {skill_title}. "
                    f"Sector: {sector}. Track: {track}. "
                    f"Level: {clean(skill.get('proficiency_level'))}. "
                    f"Description: {clean(skill.get('description'))}. "
                    f"Proficiency: {clean(skill.get('proficiency_description'))}"
                )
                yield VectorPoint(
                    id=f"joined-role-{row_index}-skill-{skill_index}",
                    document=document,
                    payload={
                        "source": source,
                        "record_type": "role_skill",
                        "role": role,
                        "skill": skill_title,
                        "sector": sector,
                        "track": track,
                        "skill_code": clean(skill.get("skill_code")),
                        "skill_type": clean(skill.get("type")),
                        "category": clean(skill.get("category")),
                        "description": clean(skill.get("description")),
                        "proficiency_level": clean(skill.get("proficiency_level")),
                        "proficiency_description": clean(skill.get("proficiency_description")),
                        "status": clean(skill.get("status")),
                    },
                )

            for function_index, function in enumerate(
                record.get("critical_work_functions") or [],
                start=1,
            ):
                critical_work_function = clean(function.get("critical_work_function"))
                for task_index, task in enumerate(function.get("key_tasks") or [], start=1):
                    task_text = clean(task)
                    if not task_text:
                        continue
                    yield VectorPoint(
                        id=f"joined-role-{row_index}-task-{function_index}-{task_index}",
                        document=(
                            f"Key task for {role}: {task_text}. "
                            f"Critical work function: {critical_work_function}"
                        ),
                        payload={
                            "source": source,
                            "record_type": "key_task",
                            "role": role,
                            "sector": sector,
                            "track": track,
                            "task": task_text,
                            "critical_work_function": critical_work_function,
                        },
                    )


def load_unique_skills(path: Path):
    if not path.exists():
        return
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Unique Skills List"]
    headers = read_headers(worksheet)
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        record = row_to_dict(headers, row)
        title = clean(record.get("skill_title"))
        description = clean(record.get("skill_description"))
        if not title:
            continue
        document = f"Unique skill: {title}. Description: {description}"
        yield VectorPoint(
            id=f"unique-skill-{row_index}",
            document=document,
            payload={
                "source": path.name,
                "record_type": "unique_skill",
                "skill": title,
                "description": description,
                "skill_type": clean(record.get("skill_type")),
                "emerging_skill": clean(record.get("Emerging Skills")),
                "casl_skill": clean(record.get("CASL Skills")),
            },
        )


def load_framework_roles(path: Path):
    if not path.exists():
        return
    workbook = load_workbook(path, read_only=True, data_only=True)
    for point in load_job_roles(path, workbook):
        yield point
    for point in load_role_skills(path, workbook):
        yield point
    for point in load_key_tasks(path, workbook):
        yield point


def load_job_roles(path: Path, workbook):
    worksheet = workbook["Job Role_Description"]
    headers = read_headers(worksheet)
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        record = row_to_dict(headers, row)
        role = clean(record.get("Job Role"))
        if not role:
            continue
        sector = clean(record.get("Sector"))
        track = clean(record.get("Track"))
        description = clean(record.get("Job Role Description"))
        document = f"Role: {role}. Sector: {sector}. Track: {track}. Description: {description}"
        yield VectorPoint(
            id=f"job-role-{row_index}",
            document=document,
            payload={
                "source": path.name,
                "record_type": "job_role",
                "role": role,
                "sector": sector,
                "track": track,
                "description": description,
            },
        )


def load_role_skills(path: Path, workbook):
    worksheet = workbook["Job Role_TCS_CCS"]
    headers = read_headers(worksheet)
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        record = row_to_dict(headers, row)
        role = clean(record.get("Job Role"))
        skill = clean(record.get("TSC_CCS Title"))
        if not role or not skill:
            continue
        sector = clean(record.get("Sector"))
        proficiency = clean(record.get("Proficiency Level"))
        document = f"Role skill: {role} requires {skill}. Sector: {sector}. Level: {proficiency}"
        yield VectorPoint(
            id=f"role-skill-{row_index}",
            document=document,
            payload={
                "source": path.name,
                "record_type": "role_skill",
                "role": role,
                "skill": skill,
                "sector": sector,
                "track": clean(record.get("Track")),
                "proficiency_level": proficiency,
                "skill_code": clean(record.get("TSC_CCS Code")),
            },
        )


def load_key_tasks(path: Path, workbook):
    worksheet = workbook["Job Role_CWF_KT"]
    headers = read_headers(worksheet)
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        record = row_to_dict(headers, row)
        role = clean(record.get("Job Role"))
        task = clean(record.get("Key Tasks"))
        if not role or not task:
            continue
        document = (
            f"Key task for {role}: {task}. Critical work function: "
            f"{clean(record.get('Critical Work Function'))}"
        )
        yield VectorPoint(
            id=f"key-task-{row_index}",
            document=document,
            payload={
                "source": path.name,
                "record_type": "key_task",
                "role": role,
                "sector": clean(record.get("Sector")),
                "track": clean(record.get("Track")),
                "task": task,
                "critical_work_function": clean(record.get("Critical Work Function")),
            },
        )


def load_tsc_mappings(path: Path):
    if not path.exists():
        return
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["data"]
    headers = read_headers(worksheet)
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        record = row_to_dict(headers, row)
        framework_skill = clean(record.get("skills_framework_skill_title"))
        unique_skill = clean(record.get("Unique skill_updated_skill_title"))
        if not framework_skill and not unique_skill:
            continue
        document = (
            f"Skills Framework skill: {framework_skill}. Unique skill: {unique_skill}. "
            f"Description: {clean(record.get('skills_framework_skill_desc'))}"
        )
        yield VectorPoint(
            id=f"tsc-unique-mapping-{row_index}",
            document=document,
            payload={
                "source": path.name,
                "record_type": "tsc_unique_mapping",
                "skill": unique_skill or framework_skill,
                "framework_skill": framework_skill,
                "unique_skill": unique_skill,
                "skill_code": clean(record.get("skills_framework_skill_code")),
                "description": clean(record.get("Unique skill_updated_skill_desc"))
                or clean(record.get("skills_framework_skill_desc")),
                "proficiency_level": clean(record.get("skills_framework_skill_pl")),
            },
        )


def read_headers(worksheet) -> list[str]:
    return [
        clean(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    ]


def row_to_dict(headers: list[str], row: tuple[object, ...]) -> dict[str, object]:
    return {
        header: row[index] if index < len(row) else None for index, header in enumerate(headers)
    }


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def fuse_matches_by_name(matches: list[VectorSearchHit], limit: int) -> list[VectorSearchHit]:
    fused: dict[str, VectorSearchHit] = {}
    for match in sorted(matches, key=lambda item: item.score, reverse=True):
        key = display_name_key(match)
        existing = fused.get(key)
        if existing is None:
            payload = dict(match.payload)
            payload["matched_ids"] = [match.id]
            payload["record_types"] = unique_clean_values([payload.get("record_type")])
            payload["roles"] = unique_clean_values([payload.get("role")])
            payload["skills"] = unique_clean_values([payload.get("skill")])
            payload["tasks"] = unique_clean_values([payload.get("task")])
            payload["descriptions"] = unique_clean_values([payload.get("description")])
            payload["documents"] = unique_clean_values([payload.get("document")])
            payload["sources"] = unique_clean_values([payload.get("source")])
            fused[key] = VectorSearchHit(id=match.id, score=match.score, payload=payload)
            continue

        existing.payload["matched_ids"] = unique_clean_values(
            [*existing.payload.get("matched_ids", []), match.id]
        )
        existing.payload["record_types"] = unique_clean_values(
            [*existing.payload.get("record_types", []), match.payload.get("record_type")]
        )
        existing.payload["roles"] = unique_clean_values(
            [*existing.payload.get("roles", []), match.payload.get("role")]
        )
        existing.payload["skills"] = unique_clean_values(
            [*existing.payload.get("skills", []), match.payload.get("skill")]
        )
        existing.payload["tasks"] = unique_clean_values(
            [*existing.payload.get("tasks", []), match.payload.get("task")]
        )
        existing.payload["descriptions"] = unique_clean_values(
            [*existing.payload.get("descriptions", []), match.payload.get("description")]
        )
        existing.payload["documents"] = unique_clean_values(
            [*existing.payload.get("documents", []), match.payload.get("document")]
        )
        existing.payload["sources"] = unique_clean_values(
            [*existing.payload.get("sources", []), match.payload.get("source")]
        )
        existing.score = max(existing.score, match.score)

    return sorted(fused.values(), key=lambda match: match.score, reverse=True)[:limit]


def display_name_key(match: VectorSearchHit) -> str:
    payload = match.payload
    name = clean(payload.get("role")) or clean(payload.get("skill")) or clean(payload.get("task"))
    if not name:
        name = match.id
    return name.casefold()


def unique_clean_values(values: list[object]) -> list[str]:
    unique_values: list[str] = []
    seen: set[str] = set()
    for value in values:
        if isinstance(value, list):
            candidates = value
        else:
            candidates = [value]

        for candidate in candidates:
            cleaned = clean(candidate)
            key = cleaned.casefold()
            if not cleaned or key in seen:
                continue
            seen.add(key)
            unique_values.append(cleaned)

    return unique_values


def normalize_tokens(text: str) -> set[str]:
    return {
        token.strip(".,:/()[]{}").lower() for token in text.split() if token.strip(".,:/()[]{}")
    }


def role_match_score(query_tokens: set[str], role: str) -> float:
    role_tokens = normalize_tokens(role)
    if not role_tokens:
        return 0

    role_text = role.lower()
    query_text = " ".join(sorted(query_tokens))
    if query_text and query_text == role_text:
        return 1.0
    if query_text and query_text in role_text:
        return 0.95

    overlap = query_tokens.intersection(role_tokens)
    if not overlap:
        return 0

    return round(len(overlap) / len(query_tokens.union(role_tokens)), 6)


def sanitize_metadata(metadata: dict[str, object]) -> dict[str, str | int | float | bool]:
    return {
        key: value for key, value in metadata.items() if isinstance(value, str | int | float | bool)
    }


def metadata_to_document(metadata: dict[str, object]) -> str:
    return ". ".join(f"{key}: {value}" for key, value in metadata.items() if value)


def resolve_backend_path(path: str) -> Path:
    candidate = Path(path)
    if candidate.is_absolute():
        return candidate
    return (BACKEND_ROOT / candidate).resolve()


vector_db_client = VectorDbClient()
